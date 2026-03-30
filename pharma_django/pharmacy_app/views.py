from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate, TruncMonth
import datetime
import json
from decimal import Decimal
from functools import wraps

from openpyxl import Workbook, load_workbook

from .models import (
    Medicine, GenericName, MedicinePresentation, Supplier,
    Purchase, Sale, SaleItem, ChatMessage, OCRPrescription,
    CustomerProfile, OnlineCart, OnlineCartItem, OnlinePrescription,
    OnlineOrder, OnlineOrderItem, OnlinePayment
)


# ─────────────────────────────────────────────────────────────
# AUTH VIEWS
# ─────────────────────────────────────────────────────────────

def _role_redirect_name(user):
    if user.is_superuser:
        return 'admin_dashboard'
    if user.is_staff:
        return 'staff_dashboard'
    return 'customer_home'


def login_view(request):
    if request.user.is_authenticated:
        return redirect(_role_redirect_name(request.user))

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        if not User.objects.filter(username=username).exists():
            messages.error(request, 'You are a new user. Please sign up first.')
            return render(request, 'registration/login.html', {
                'active_tab': 'signup',
                'show_signup_popup': True,
            })
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}.')
            return redirect(_role_redirect_name(user))
        messages.error(request, 'Invalid username or password.')
        return render(request, 'registration/login.html', {'active_tab': 'login'})

    return render(request, 'registration/login.html', {'active_tab': 'login'})


def signup_view(request):
    if request.user.is_authenticated:
        return redirect(_role_redirect_name(request.user))

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        role = request.POST.get('role', 'customer').strip()

        if not username or not password:
            messages.error(request, 'Username and password are required for signup.')
            return render(request, 'registration/login.html', {'active_tab': 'signup'})
        if password != confirm_password:
            messages.error(request, 'Password and confirm password must match.')
            return render(request, 'registration/login.html', {'active_tab': 'signup'})
        if len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
            return render(request, 'registration/login.html', {'active_tab': 'signup'})
        if User.objects.filter(username=username).exists():
            messages.error(request, 'This username is already taken.')
            return render(request, 'registration/login.html', {'active_tab': 'signup'})
        if email and User.objects.filter(email__iexact=email).exists():
            messages.error(request, 'This email is already registered.')
            return render(request, 'registration/login.html', {'active_tab': 'signup'})

        # Frontend signup is customer-only by design.
        if role != 'customer':
            messages.error(request, 'Admin roles can only be assigned by Super Admin.')
            return render(request, 'registration/login.html', {'active_tab': 'signup'})

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            is_staff=False,
            is_superuser=False,
        )
        if full_name:
            name_parts = full_name.split()
            user.first_name = name_parts[0]
            user.last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
            user.save(update_fields=['first_name', 'last_name'])

        CustomerProfile.objects.get_or_create(user=user)
        messages.success(request, 'Signup successful. Please login with your account.')
        return redirect('login')

    return render(request, 'registration/login.html', {'active_tab': 'signup'})


def logout_view(request):
    logout(request)
    return redirect('login')


def staff_required(view_func):
    """Allow only super admin and staff admin users."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_staff:
            messages.error(request, 'This section is available for pharmacy admin users only.')
            return redirect('customer_home')
        return view_func(request, *args, **kwargs)
    return _wrapped


def superadmin_required(view_func):
    """Allow only super admin users."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_superuser:
            messages.error(request, 'This section is available for Super Admin only.')
            return redirect('staff_dashboard' if request.user.is_staff else 'customer_home')
        return view_func(request, *args, **kwargs)
    return _wrapped


@login_required
def role_home(request):
    return redirect(_role_redirect_name(request.user))


@login_required
@superadmin_required
def admin_dashboard(request):
    context = {
        'total_users': User.objects.count(),
        'total_super_admins': User.objects.filter(is_superuser=True, is_staff=True).count(),
        'total_admin_staff': User.objects.filter(is_staff=True, is_superuser=False).count(),
        'total_customers': User.objects.filter(is_staff=False, is_superuser=False).count(),
        'recent_users': User.objects.order_by('-date_joined')[:10],
    }
    return render(request, 'pharmacy_app/admin_dashboard.html', context)


# ─────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    if not request.user.is_staff:
        return redirect('customer_home')

    today = timezone.now().date()
    month_start = today.replace(day=1)

    # Key stats
    total_medicines = Medicine.objects.count()
    out_of_stock = Medicine.objects.filter(quantity=0).count()
    low_stock = Medicine.objects.filter(quantity__gt=0, quantity__lt=20).count()
    expiring_soon = Medicine.objects.filter(
        expire_date__lte=today + datetime.timedelta(days=30),
        expire_date__gte=today
    ).count()
    expired = Medicine.objects.filter(expire_date__lt=today).count()
    total_suppliers = Supplier.objects.count()
    total_generics = GenericName.objects.count()

    today_sales = Sale.objects.filter(
        sale_date__date=today
    ).aggregate(total=Sum('sub_total'))['total'] or 0

    month_sales = Sale.objects.filter(
        sale_date__date__gte=month_start
    ).aggregate(total=Sum('sub_total'))['total'] or 0

    total_sales = Sale.objects.aggregate(total=Sum('sub_total'))['total'] or 0
    total_online_orders = OnlineOrder.objects.count()
    pending_online_orders = OnlineOrder.objects.filter(status='pending').count()
    online_revenue = OnlineOrder.objects.filter(
        status__in=['approved', 'packed', 'out_for_delivery', 'delivered']
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Sales chart — last 14 days
    sales_by_day = (
        Sale.objects
        .filter(sale_date__date__gte=today - datetime.timedelta(days=13))
        .annotate(day=TruncDate('sale_date'))
        .values('day')
        .annotate(revenue=Sum('sub_total'), count=Count('id'))
        .order_by('day')
    )
    chart_labels = []
    chart_data = []
    for i in range(14):
        d = today - datetime.timedelta(days=13 - i)
        chart_labels.append(d.strftime('%b %d'))
        rev = next((x['revenue'] for x in sales_by_day if x['day'] == d), 0)
        chart_data.append(float(rev or 0))

    # Top 5 selling medicines
    top_medicines = (
        SaleItem.objects
        .values('medicine__name')
        .annotate(total_sold=Sum('quantity'), revenue=Sum('sub_total'))
        .order_by('-total_sold')[:5]
    )

    # Recent 8 sales
    recent_sales = Sale.objects.select_related('created_by').prefetch_related('items__medicine')[:8]

    # Stock distribution for donut chart
    stock_dist = {
        'Out of Stock': out_of_stock,
        'Low Stock': low_stock,
        'Adequate': Medicine.objects.filter(quantity__gte=20, quantity__lt=100).count(),
        'Well Stocked': Medicine.objects.filter(quantity__gte=100).count(),
    }

    # Alerts
    alerts = Medicine.objects.filter(
        Q(quantity__lt=20) | Q(expire_date__lte=today + datetime.timedelta(days=30))
    ).order_by('quantity')[:8]

    dashboard_details = {
        'total_medicines': {
            'title': 'All Medicines',
            'items': list(
                Medicine.objects.select_related('generic_name', 'presentation', 'supplier')
                .values('name', 'generic_name__name', 'presentation__name', 'quantity', 'selling_price')
                .order_by('name')[:50]
            ),
        },
        'today_sales': {
            'title': "Today's Sales",
            'items': list(
                Sale.objects.filter(sale_date__date=today)
                .select_related('created_by')
                .values('invoice_number', 'customer_name', 'sub_total', 'sale_date', 'created_by__username')
                .order_by('-sale_date')[:30]
            ),
        },
        'out_of_stock': {
            'title': 'Out of Stock Medicines',
            'items': list(
                Medicine.objects.filter(quantity=0)
                .select_related('generic_name', 'presentation')
                .values('name', 'generic_name__name', 'presentation__name', 'expire_date')
                .order_by('name')[:50]
            ),
        },
        'expiring_soon': {
            'title': 'Expiring Within 30 Days',
            'items': list(
                Medicine.objects.filter(
                    expire_date__lte=today + datetime.timedelta(days=30),
                    expire_date__gte=today
                )
                .select_related('generic_name', 'presentation')
                .values('name', 'generic_name__name', 'presentation__name', 'quantity', 'expire_date')
                .order_by('expire_date', 'name')[:50]
            ),
        },
        'suppliers': {
            'title': 'Active Suppliers',
            'items': list(
                Supplier.objects.annotate(medicine_count=Count('medicine'))
                .values('company_name', 'contact_person', 'phone', 'previous_due', 'medicine_count')
                .order_by('company_name')[:50]
            ),
        },
        'generics': {
            'title': 'Generic Names',
            'items': list(
                GenericName.objects.annotate(medicine_count=Count('medicines'))
                .values('name', 'description', 'medicine_count')
                .order_by('name')[:50]
            ),
        },
        'total_sales': {
            'title': 'All-Time Sales',
            'items': list(
                Sale.objects.select_related('created_by')
                .values('invoice_number', 'customer_name', 'sub_total', 'sale_date', 'created_by__username')
                .order_by('-sale_date')[:50]
            ),
        },
        'top_medicines': {
            'title': 'Top Selling Medicines',
            'items': list(top_medicines),
        },
        'alerts': {
            'title': 'Stock And Expiry Alerts',
            'items': list(
                alerts.values('name', 'quantity', 'expire_date')
            ),
        },
    }

    context = {
        'total_medicines': total_medicines,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'expiring_soon': expiring_soon,
        'expired': expired,
        'total_suppliers': total_suppliers,
        'total_generics': total_generics,
        'today_sales': today_sales,
        'month_sales': month_sales,
        'total_sales': total_sales,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'top_medicines': json.dumps(list(top_medicines), default=str),
        'stock_dist': json.dumps(stock_dist),
        'recent_sales': recent_sales,
        'alerts': alerts,
        'dashboard_details': json.dumps(dashboard_details, default=str),
        'today': today,
        'total_online_orders': total_online_orders,
        'pending_online_orders': pending_online_orders,
        'online_revenue': online_revenue,
    }
    return render(request, 'pharmacy_app/dashboard.html', context)


@login_required
@staff_required
def staff_dashboard(request):
    return dashboard(request)


@login_required
def customer_home(request):
    if request.user.is_staff:
        return redirect(_role_redirect_name(request.user))
    return render(request, 'pharmacy_app/customer_home.html')


# ─────────────────────────────────────────────────────────────
# MEDICINES
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def medicines_view(request):
    generics = GenericName.objects.all()
    presentations = MedicinePresentation.objects.all()
    suppliers = Supplier.objects.all()
    context = {
        'generics': generics,
        'presentations': presentations,
        'suppliers': suppliers,
    }
    return render(request, 'pharmacy_app/medicines.html', context)


# ─────────────────────────────────────────────────────────────
# GENERIC NAMES
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def generics_view(request):
    return render(request, 'pharmacy_app/generics.html')


# ─────────────────────────────────────────────────────────────
# SUPPLIERS
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def suppliers_view(request):
    return render(request, 'pharmacy_app/suppliers.html')


# ─────────────────────────────────────────────────────────────
# INVENTORY / PURCHASE
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def inventory_view(request):
    medicines_list = Medicine.objects.select_related('generic_name', 'presentation', 'supplier').all()
    suppliers = Supplier.objects.all()
    generics = GenericName.objects.all()
    presentations = MedicinePresentation.objects.all()
    context = {
        'medicines_list': medicines_list,
        'suppliers': suppliers,
        'generics': generics,
        'presentations': presentations,
    }
    return render(request, 'pharmacy_app/inventory.html', context)


# ─────────────────────────────────────────────────────────────
# SALES
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def sales_view(request):
    medicines_qs = Medicine.objects.filter(quantity__gt=0).values('id', 'name', 'selling_price', 'quantity', 'volume')
    context = {
        'medicines_json': json.dumps(list(medicines_qs), default=str),
    }
    return render(request, 'pharmacy_app/sales.html', context)


@login_required
@staff_required
def sales_list_view(request):
    return render(request, 'pharmacy_app/sales_list.html')


# ─────────────────────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def reports_view(request):
    return render(request, 'pharmacy_app/reports.html')


# ─────────────────────────────────────────────────────────────
# OCR PRESCRIPTION
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def ocr_view(request):
    recent = OCRPrescription.objects.filter(uploaded_by=request.user)[:5]
    return render(request, 'pharmacy_app/ocr.html', {'recent': recent})


# ─────────────────────────────────────────────────────────────
# RECOMMEND
# ─────────────────────────────────────────────────────────────

@login_required
@staff_required
def recommend_view(request):
    return render(request, 'pharmacy_app/recommend.html')


@login_required
@staff_required
def online_purchase_view(request):
    selected_status = request.GET.get('status', '').strip()
    online_orders_qs = OnlineOrder.objects.select_related('customer__user').all()
    if selected_status:
        online_orders_qs = online_orders_qs.filter(status=selected_status)

    recent_online_orders = online_orders_qs[:12]
    pending_prescriptions = OnlinePrescription.objects.select_related('customer__user').filter(status='pending')[:10]
    total_online_orders = OnlineOrder.objects.count()
    pending_online_orders = OnlineOrder.objects.filter(status='pending').count()
    delivered_online_orders = OnlineOrder.objects.filter(status='delivered').count()
    online_revenue = OnlineOrder.objects.filter(
        status__in=['approved', 'packed', 'out_for_delivery', 'delivered']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    pending_prescription_count = OnlinePrescription.objects.filter(status='pending').count()

    context = {
        'order_status_choices': OnlineOrder.STATUS_CHOICES,
        'selected_status': selected_status,
        'recent_online_orders': recent_online_orders,
        'pending_prescriptions': pending_prescriptions,
        'total_online_orders': total_online_orders,
        'pending_online_orders': pending_online_orders,
        'delivered_online_orders': delivered_online_orders,
        'online_revenue': online_revenue,
        'pending_prescription_count': pending_prescription_count,
    }
    return render(request, 'pharmacy_app/online_purchase.html', context)


def _get_customer_profile(user):
    profile, _ = CustomerProfile.objects.get_or_create(user=user)
    if not profile.phone:
        profile.phone = ''
        profile.save(update_fields=['phone'])
    return profile


def _get_online_cart(user):
    profile = _get_customer_profile(user)
    cart, _ = OnlineCart.objects.get_or_create(customer=profile)
    return profile, cart


@login_required
def online_medicines_view(request):
    query = request.GET.get('q', '').strip()
    medicines = Medicine.objects.filter(quantity__gt=0, is_online_available=True)
    if query:
        medicines = medicines.filter(
            Q(name__icontains=query) |
            Q(generic_name__name__icontains=query) |
            Q(description__icontains=query)
        )
    medicines = medicines.select_related('generic_name', 'presentation').order_by('name')
    return render(request, 'pharmacy_app/online_medicines.html', {'medicines': medicines, 'query': query})


@login_required
def online_cart_view(request):
    profile, cart = _get_online_cart(request.user)
    items = cart.items.select_related('medicine').order_by('-added_at')
    total = sum((item.line_total for item in items), Decimal('0'))
    return render(request, 'pharmacy_app/online_cart.html', {
        'profile': profile,
        'cart': cart,
        'items': items,
        'total': total,
    })


@login_required
def online_add_to_cart(request, medicine_id):
    medicine = get_object_or_404(Medicine, id=medicine_id)
    if medicine.quantity <= 0:
        messages.error(request, f'{medicine.name} is out of stock.')
        return redirect('online_medicines')

    _, cart = _get_online_cart(request.user)
    item, created = OnlineCartItem.objects.get_or_create(cart=cart, medicine=medicine, defaults={'quantity': 1})
    if not created:
        if item.quantity + 1 > medicine.quantity:
            messages.error(request, 'Requested quantity exceeds stock.')
            return redirect('online_cart')
        item.quantity += 1
        item.save(update_fields=['quantity'])
    messages.success(request, f'{medicine.name} added to cart.')
    return redirect('online_cart')


@login_required
def online_update_cart_item(request, item_id):
    item = get_object_or_404(OnlineCartItem, id=item_id, cart__customer__user=request.user)
    if request.method != 'POST':
        return redirect('online_cart')

    try:
        quantity = int(request.POST.get('quantity', '1'))
    except ValueError:
        quantity = 1

    if quantity <= 0:
        item.delete()
        messages.success(request, 'Item removed from cart.')
        return redirect('online_cart')

    if quantity > item.medicine.quantity:
        messages.error(request, f'Only {item.medicine.quantity} units available for {item.medicine.name}.')
        return redirect('online_cart')

    item.quantity = quantity
    item.save(update_fields=['quantity'])
    messages.success(request, 'Cart updated.')
    return redirect('online_cart')


@login_required
def online_remove_cart_item(request, item_id):
    item = get_object_or_404(OnlineCartItem, id=item_id, cart__customer__user=request.user)
    item.delete()
    messages.success(request, 'Item removed from cart.')
    return redirect('online_cart')


@login_required
def online_upload_prescription(request):
    profile = _get_customer_profile(request.user)
    if request.method == 'POST' and request.FILES.get('image'):
        OnlinePrescription.objects.create(
            customer=profile,
            image=request.FILES['image'],
            status='pending'
        )
        messages.success(request, 'Prescription uploaded. Please wait for approval.')
    else:
        messages.error(request, 'Please choose a prescription image to upload.')
    return redirect('online_checkout')


@login_required
@transaction.atomic
def online_checkout_view(request):
    profile, cart = _get_online_cart(request.user)
    items = list(cart.items.select_related('medicine').order_by('-added_at'))
    if not items:
        messages.error(request, 'Your cart is empty.')
        return redirect('online_medicines')

    cart_total = sum((item.line_total for item in items), Decimal('0'))
    needs_prescription = any(item.medicine.prescription_required for item in items)
    selected_prescription_id = request.POST.get('prescription_id') if request.method == 'POST' else ''
    prescriptions = profile.online_prescriptions.all()[:10]

    if request.method == 'POST':
        customer_name = request.POST.get('customer_name', '').strip() or request.user.get_full_name() or request.user.username
        customer_phone = request.POST.get('customer_phone', '').strip() or profile.phone
        delivery_address = request.POST.get('delivery_address', '').strip() or profile.full_address
        notes = request.POST.get('notes', '').strip()
        payment_method = request.POST.get('payment_method', 'cod')

        if not customer_phone or not delivery_address:
            messages.error(request, 'Phone and delivery address are required.')
            return redirect('online_checkout')

        prescription_obj = None
        if needs_prescription:
            if not selected_prescription_id:
                messages.error(request, 'Prescription is required for restricted medicines.')
                return redirect('online_checkout')
            prescription_obj = get_object_or_404(
                OnlinePrescription,
                id=selected_prescription_id,
                customer=profile
            )
            if prescription_obj.status != 'approved':
                messages.error(request, 'Selected prescription is not approved yet.')
                return redirect('online_checkout')

        medicine_ids = [item.medicine_id for item in items]
        locked_medicines = {
            med.id: med for med in Medicine.objects.select_for_update().filter(id__in=medicine_ids)
        }
        for item in items:
            medicine = locked_medicines.get(item.medicine_id)
            if not medicine or medicine.quantity < item.quantity:
                messages.error(request, f'Insufficient stock for {item.medicine.name}.')
                return redirect('online_cart')

        order = OnlineOrder.objects.create(
            customer=profile,
            prescription=prescription_obj,
            status='pending',
            payment_method=payment_method,
            customer_name=customer_name,
            customer_phone=customer_phone,
            delivery_address=delivery_address,
            notes=notes,
            total_amount=cart_total,
        )
        order.log_status('pending', changed_by=request.user, note='Order placed from web checkout')

        for item in items:
            medicine = locked_medicines[item.medicine_id]
            line_total = item.quantity * medicine.selling_price
            OnlineOrderItem.objects.create(
                order=order,
                medicine=medicine,
                quantity=item.quantity,
                unit_price=medicine.selling_price,
                line_total=line_total
            )
            medicine.quantity -= item.quantity
            medicine.save(update_fields=['quantity'])

        OnlinePayment.objects.create(
            order=order,
            method=payment_method,
            status='pending',
            gateway='cod' if payment_method == 'cod' else 'online',
            amount=cart_total,
        )
        cart.items.all().delete()
        messages.success(request, f'Order placed successfully. Order ID: {order.order_id}')
        return redirect('online_order_detail', order_id=order.id)

    return render(request, 'pharmacy_app/online_checkout.html', {
        'profile': profile,
        'items': items,
        'cart_total': cart_total,
        'needs_prescription': needs_prescription,
        'prescriptions': prescriptions,
        'selected_prescription_id': selected_prescription_id,
    })


@login_required
def online_orders_view(request):
    status_filter = request.GET.get('status', '').strip()
    orders = OnlineOrder.objects.select_related('customer__user').prefetch_related('items__medicine')
    if status_filter:
        orders = orders.filter(status=status_filter)
    if not request.user.is_staff:
        orders = orders.filter(customer__user=request.user)
    return render(request, 'pharmacy_app/online_orders.html', {
        'orders': orders[:100],
        'status_filter': status_filter,
        'status_choices': OnlineOrder.STATUS_CHOICES,
    })


@login_required
def online_order_detail_view(request, order_id):
    order = get_object_or_404(
        OnlineOrder.objects.select_related('customer__user', 'prescription', 'payment').prefetch_related('items__medicine'),
        id=order_id
    )
    if not request.user.is_staff and order.customer.user_id != request.user.id:
        messages.error(request, 'You are not allowed to view this order.')
        return redirect('online_orders')
    payment = getattr(order, 'payment', None)
    return render(request, 'pharmacy_app/online_order_detail.html', {'order': order, 'payment': payment})


@login_required
def online_update_order_status(request, order_id):
    if request.method != 'POST':
        return redirect('online_orders')
    if not request.user.is_staff:
        messages.error(request, 'Only admin/staff can update order status.')
        return redirect('online_orders')

    order = get_object_or_404(OnlineOrder, id=order_id)
    new_status = request.POST.get('status', '').strip()
    valid_statuses = {choice[0] for choice in OnlineOrder.STATUS_CHOICES}
    if new_status not in valid_statuses:
        messages.error(request, 'Invalid order status.')
        return redirect('online_orders')
    order.status = new_status
    order.save(update_fields=['status', 'updated_at'])
    order.log_status(new_status, changed_by=request.user, note='Updated from web order screen')
    messages.success(request, f'Order {order.order_id} status updated to {order.get_status_display()}.')
    return redirect('online_orders')


@login_required
def online_review_prescription(request, prescription_id):
    if request.method != 'POST':
        return redirect('online_purchase')
    if not request.user.is_staff:
        messages.error(request, 'Only admin/staff can review prescriptions.')
        return redirect('online_purchase')

    prescription = get_object_or_404(OnlinePrescription, id=prescription_id)
    action = request.POST.get('action', '').strip()
    note = request.POST.get('review_note', '').strip()
    if action not in {'approved', 'rejected'}:
        messages.error(request, 'Invalid prescription action.')
        return redirect('online_purchase')

    prescription.status = action
    prescription.review_note = note
    prescription.reviewed_by = request.user
    prescription.reviewed_at = timezone.now()
    prescription.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at'])
    messages.success(request, f'Prescription #{prescription.id} marked as {action}.')
    return redirect('online_purchase')


@login_required
def online_invoice_view(request, order_id):
    order = get_object_or_404(
        OnlineOrder.objects.select_related('customer__user').prefetch_related('items__medicine'),
        id=order_id
    )
    if not request.user.is_staff and order.customer.user_id != request.user.id:
        messages.error(request, 'You are not allowed to access this invoice.')
        return redirect('online_orders')
    return render(request, 'pharmacy_app/online_invoice.html', {'order': order})


@login_required
@staff_required
def excel_tools_view(request):
    return render(request, 'pharmacy_app/excel_tools.html')


@login_required
@staff_required
def export_medicines_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Medicines'
    sheet.append([
        'name', 'generic_name', 'presentation', 'volume', 'unit_price',
        'selling_price', 'quantity', 'expire_date', 'supplier', 'description'
    ])

    medicines = Medicine.objects.select_related('generic_name', 'presentation', 'supplier').order_by('name')
    for medicine in medicines:
        sheet.append([
            medicine.name,
            medicine.generic_name.name if medicine.generic_name else '',
            medicine.presentation.name if medicine.presentation else '',
            medicine.volume,
            float(medicine.unit_price),
            float(medicine.selling_price),
            medicine.quantity,
            medicine.expire_date.isoformat() if medicine.expire_date else '',
            medicine.supplier.company_name if medicine.supplier else '',
            medicine.description,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pharmacy_medicines.xlsx'
    workbook.save(response)
    return response


@login_required
@staff_required
def import_medicines_excel(request):
    if request.method != 'POST':
        return redirect('excel_tools')

    upload = request.FILES.get('excel_file')
    if not upload:
        messages.error(request, 'Please choose an Excel file first.')
        return redirect('excel_tools')

    try:
        workbook = load_workbook(upload)
        sheet = workbook.active
        header = [str(cell.value).strip() if cell.value is not None else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    except Exception:
        messages.error(request, 'Could not read the Excel file. Please upload a valid .xlsx file.')
        return redirect('excel_tools')

    required_columns = {
        'name', 'generic_name', 'presentation', 'volume', 'unit_price',
        'selling_price', 'quantity', 'expire_date', 'supplier', 'description'
    }
    if not required_columns.issubset(set(header)):
        messages.error(request, 'Excel columns are invalid. Export the sample file first and use the same format.')
        return redirect('excel_tools')

    index = {column: position for position, column in enumerate(header)}
    created_count = 0
    updated_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[index['name']]).strip() if row[index['name']] is not None else ''
        if not name:
            continue

        generic_name_value = str(row[index['generic_name']]).strip() if row[index['generic_name']] is not None else ''
        presentation_value = str(row[index['presentation']]).strip() if row[index['presentation']] is not None else ''
        supplier_value = str(row[index['supplier']]).strip() if row[index['supplier']] is not None else ''
        expire_value = row[index['expire_date']]

        generic = None
        if generic_name_value:
            generic, _ = GenericName.objects.get_or_create(name=generic_name_value)

        presentation = None
        if presentation_value:
            presentation, _ = MedicinePresentation.objects.get_or_create(name=presentation_value)

        supplier = None
        if supplier_value:
            supplier, _ = Supplier.objects.get_or_create(company_name=supplier_value)

        expire_date = None
        if expire_value:
            if isinstance(expire_value, datetime.datetime):
                expire_date = expire_value.date()
            elif isinstance(expire_value, datetime.date):
                expire_date = expire_value
            else:
                try:
                    expire_date = datetime.date.fromisoformat(str(expire_value))
                except ValueError:
                    expire_date = None

        medicine, created = Medicine.objects.update_or_create(
            name=name,
            defaults={
                'generic_name': generic,
                'presentation': presentation,
                'volume': str(row[index['volume']] or '').strip(),
                'unit_price': Decimal(str(row[index['unit_price']] or 0)),
                'selling_price': Decimal(str(row[index['selling_price']] or 0)),
                'quantity': int(row[index['quantity']] or 0),
                'expire_date': expire_date,
                'supplier': supplier,
                'description': str(row[index['description']] or '').strip(),
            }
        )

        if created:
            created_count += 1
        else:
            updated_count += 1

    messages.success(request, f'Excel import complete. Created: {created_count}, Updated: {updated_count}.')
    return redirect('excel_tools')
